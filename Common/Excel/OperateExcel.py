#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
Created on 2016年8月21日
@author: LiuBin
@note: excel operation
"""

from openpyxl.reader.excel import load_workbook
# from PythonAW.LoggingAW.LogAW import logger


class OperateExcel():
    def __init__(self):
        self.lw_aw = None
        pass

    def get_excel_max_row(self, file_name):
        """
            - File: excel path. ex: c:\\webtc\\textopenpyxl.xlsx \n
        Return:\n
            - ws: excel max number \n
        Author: LiuBin  Date: 2018-11-13 \n
        Notes: Gets the maximum number of rows in the excel table. \n
        Examples: ${value} | GetExcelMaxRow | webtc.xlsx \n
        """
        self.lw_aw = load_workbook(file_name)
        sheet_names = self.lw_aw.get_sheet_names()
        ws = self.lw_aw.get_sheet_by_name(sheet_names[0])
        # logger.info("the result is " + str(ws.max_row))
        return ws.max_row

    def get_excel_info_by_cell(self, file_name, row_num, col_mum):
        """
            - File: excel path. ex: c:\\webtc\\textopenpyxl.xlsx \n
            - RowNum: row number  \n
            - ColNum: col number \n
        Return:\n
            - ws: excel value \n
        Author: LiuBin    Date: 2018-11-13 \n
        Notes: Gets value in the excel table \n
        Examples: ${value} | GetExcelInfoByCell | webtc.xlsx | 1 | 1 \n
        """
        self.lw_aw = load_workbook(file_name)
        sheet_names = self.lw_aw.get_sheet_names()
        ws = self.lw_aw.get_sheet_by_name(sheet_names[0])
        row_num = int(row_num)
        col_mum = int(col_mum)
        # logger.info("the result is " + ws.cell(row=row_num, column=col_mum).value)
        return ws.cell(row=row_num, column=col_mum).value

    def get_lun_by_feature(self, excel_path, feature):
        row = self.get_excel_max_row(excel_path)
        for index in range(1, row):
            get_feature = self.get_excel_info_by_cell(excel_path, str(index), 1)
            if get_feature in feature:
                return self.get_excel_info_by_cell(excel_path, str(index+1), 2)

    def get_lun_number(self, excel_path, feature):
        row = self.get_excel_max_row(excel_path)
        for index in range(1, row):
            get_feature = self.get_excel_info_by_cell(excel_path, str(index), 1)
            if get_feature in feature:
                bb = bb +1
        return  bb


if __name__ == "__main__":
    op_aw = OperateExcel()
    #print op_aw.get_excel_max_row("c:\\textopenpyxl.xlsx")
    #print op_aw.get_excel_info_by_cell("c:\\textopenpyxl.xlsx",1,1)
    #print op_aw.get_excel_info_by_cell("c:\\textopenpyxl.xlsx",1,2)
    print(op_aw.get_lun_by_feature("d:\\textopenpyxl.xlsx","SanHpyerAS"))
    